#! python
# Usage: python3 textToExcel.py

import openpyxl
import sys
import os
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

wb = openpyxl.Workbook()
ws = wb.active

p = Path.cwd()
text_files = list(p.glob("*.txt"))

filelines = [] # this list will be a list of a list of lines in a file

# appending lines to the lines list
for textfile in text_files:
    working_file = open(textfile, "r")
    filelines.append(working_file.readlines())

# writing lines to worksheet
for i in range(len(filelines)):
    for j in range(len(filelines[i])):
        ws.cell(row=j+1, column=i+1).value = filelines[i][j]

wb.save("lines.xlsx")