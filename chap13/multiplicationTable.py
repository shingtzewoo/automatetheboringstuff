#! python
# Usage: python3 multiplicationTable.py N
# Takes a number from command line argument and creates a multiplication table

import openpyxl
import sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

# check whether arguments equal to 2 if not, sys.exit
if len(sys.argv) != 2:
    print("Usage: python3 multiplicationTable.py N")
    sys.exit()

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Multiplication Tbl %s" % sys.argv[1]
n = int(sys.argv[1])

bold = Font(name='Times New Roman', bold=True)

# populate first row and column up to N and bold it
for i in range(1, n+1):
    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=1, column=i+1).font = bold
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=1).font = bold

# loop through cells per row, which will require two loops: first loop goes through columns, inner loop goes through row
# for each cell, multiply it by the appropriate header cell

for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        sheet.cell(row=i, column=j).value = int(sheet.cell(row=i, column=1).value) * int(sheet.cell(row=1, column=j).value)

filename = f"MultiplicationTable{sys.argv[1]}.xlsx"
wb.save(filename)